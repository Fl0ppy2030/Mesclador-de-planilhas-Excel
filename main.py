import pandas as pd
import os
from tkinter import Tk, filedialog, messagebox, Label, Button, Frame
import tkinter as tk
from openpyxl import load_workbook
from datetime import datetime


def safe_int(valor, padrao=1):
    """Converte valor para int com segurança"""
    try:
        if pd.isna(valor):
            return padrao
        return int(float(valor))
    except (ValueError, TypeError):
        return padrao


class ExcelDataMerger:
    def __init__(self):
        self.root = Tk()
        self.root.title("Mesclador de Guias de Remessa")
        self.root.geometry("550x550")

        self.source_file = None
        self.target_file = None

        self.setup_ui()

    def setup_ui(self):
        Label(self.root, text="Preenchimento de Guias de Remessa",
              font=("Arial", 14, "bold")).pack(pady=15)

        instructions = Label(
            self.root,
            text=(
                "INSTRUÇÕES:\n"
                "1. Selecione a planilha com os dados (fonte)\n"
                "2. Selecione o modelo em branco da guia (destino)\n"
                "3. Configure as linhas disponíveis\n"
                "4. Clique em Processar\n"
                "5. Os campos última linha e linhas por páginas deveram ser ajustada conforme a necessidade\n\n"
                "Dados extraídos: UNIQUEFILENAME e COD_DOCTO"
            ),
            justify="left",
            font=("Arial", 9)
        )
        instructions.pack()

        file_frame = Frame(self.root)
        file_frame.pack(pady=15)

        Label(file_frame, text="Planilha Fonte:").grid(row=0, column=0, sticky="w")
        Button(file_frame, text="Selecionar", command=self.select_source_file).grid(row=0, column=1)
        self.source_label = Label(file_frame, text="Não selecionado", fg="gray", width=30, anchor="w")
        self.source_label.grid(row=0, column=2)

        Label(file_frame, text="Modelo Destino:").grid(row=1, column=0, sticky="w")
        Button(file_frame, text="Selecionar", command=self.select_target_file).grid(row=1, column=1)
        self.target_label = Label(file_frame, text="Não selecionado", fg="gray", width=30, anchor="w")
        self.target_label.grid(row=1, column=2)

        config = Frame(self.root)
        config.pack(pady=15)

        Label(config, text="Primeira linha").grid(row=0, column=0)
        self.start_row_var = tk.StringVar(value="10")
        tk.Entry(config, textvariable=self.start_row_var, width=5).grid(row=0, column=1)

        Label(config, text="Última linha").grid(row=0, column=2)
        self.end_row_var = tk.StringVar(value="27")
        tk.Entry(config, textvariable=self.end_row_var, width=5).grid(row=0, column=3)

        Label(config, text="Linhas por página").grid(row=1, column=0)
        self.rows_per_page_var = tk.StringVar(value="18")
        tk.Entry(config, textvariable=self.rows_per_page_var, width=5).grid(row=1, column=1)

        Label(config, text="Data").grid(row=1, column=2)
        self.date_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))
        tk.Entry(config, textvariable=self.date_var, width=10).grid(row=1, column=3)

        Button(
            self.root,
            text="PROCESSAR",
            command=self.process_files,
            bg="#0066cc",
            fg="white",
            width=20,
            height=2
        ).pack(pady=15)

        self.status_label = Label(self.root, text="Aguardando seleção de arquivos...", fg="gray")
        self.status_label.pack()

    def select_source_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls *xlsm")])
        if path:
            self.source_file = path
            self.source_label.config(text=os.path.basename(path), fg="blue")

    def select_target_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls *xlsm")])
        if path:
            self.target_file = path
            self.target_label.config(text=os.path.basename(path), fg="blue")

    def extract_data_from_source(self):
        try:
            df = pd.read_excel(self.source_file)

            def find_col(names):
                for c in df.columns:
                    if any(n in str(c).upper() for n in names):
                        return c
                return None

            col_n1710 = find_col(["UNIQUEFILENAME", "N-1710"])
            col_replan = find_col(["COD_DOCTO", "REPLAN"])
            col_fl = find_col(["FL", "FOLHA"])
            col_tot = find_col(["TOT.FLS", "TOTAL_FOLHAS"])
            col_rev = find_col(["REV", "REVISAO"])

            if not col_n1710 or not col_replan:
                messagebox.showerror("Erro", "Colunas principais não encontradas.")
                return None

            dados = []

            for _, row in df.iterrows():
                n1710 = str(row[col_n1710]).strip()
                replan = str(row[col_replan]).strip()

                # IGNORA linhas inválidas / cabeçalhos
                if (
                    not n1710
                    or n1710.lower() == "nan"
                    or "lista de documentos" in n1710.lower()
                    or "nome do arquivo" in n1710.lower()
                    or "uniquefilename" in n1710.lower()
                    or not (n1710.lower().endswith(".pdf") or n1710.lower().endswith(".zip"))
                ):
                    continue

                if not replan or replan.lower() == "nan" or "cod" in replan.lower():
                    continue


                fl = safe_int(row[col_fl])
                tot = safe_int(row[col_tot])
                rev = str(row[col_rev]).strip().strip()
                dados.append([n1710, replan, fl, tot, rev, "", ""])

            return dados

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao extrair dados:\n{e}")
            return None

    def fill_target_spreadsheet(self, dados):
        wb = load_workbook(self.target_file)
        ws_base = wb.active

        start = int(self.start_row_var.get())
        end = int(self.end_row_var.get())
        per_page = int(self.rows_per_page_var.get())

        pages = (len(dados) + per_page - 1) // per_page

        for p in range(pages):
            ws = ws_base if p == 0 else wb.copy_worksheet(ws_base)
            ws.title = f"Página_{p+1}"

            row = start
            for d in dados[p*per_page:(p+1)*per_page]:
                ws.cell(row=row, column=1, value=row-start+1)
                for c in range(7):
                    ws.cell(row=row, column=c+2, value=d[c])
                row += 1

            for r in range(row, end+1):
                for c in range(1, 9):
                    ws.cell(row=r, column=c, value="")

        output = os.path.splitext(self.target_file)[0] + "_PREENC.xlsx"
        wb.save(output)
        return output

    def process_files(self):
        if not self.source_file or not self.target_file:
            messagebox.showwarning("Atenção", "Selecione os dois arquivos.")
            return

        dados = self.extract_data_from_source()
        if not dados:
            return

        output = self.fill_target_spreadsheet(dados)
        messagebox.showinfo("Sucesso", f"Arquivo gerado:\n{output}")

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    ExcelDataMerger().run()
